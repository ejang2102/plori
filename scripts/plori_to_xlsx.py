"""Export PLoRI ploridata.npz bundles to a human-readable Excel (.xlsx) workbook (for sharing with non-developers).

Combines the {cat}_{name}_ploridata.npz files written by plori_batch.py into a single unified workbook.
Transfers the three signals and computed metrics (the same ones plotted by the report card) straight into tables.

Sheet layout:
  README        - Description of the signals and metrics
  Summary       - One-line-per-sample summary (all scalar metrics)
  TS_<name>     - Per-frame time series of the three signals (PLoRI / frame_diff / dPLoRI_dt) + time_s
  beats_<name>  - Per-beat metrics (interval, CD50, CT, RT, amplitude, onset/offset;
                  with --with-flow also max speed / displacement / strain)

frame_diff (the middle panel of the report card) is not stored as a finished array in the npz, so it is
reconstructed from pixraw (plori.core.derive); caches without pixraw leave that column empty.

Usage:
  python scripts/plori_to_xlsx.py \
    --glob 'results/**/*_ploridata.npz' \
    --out  results/PLoRI_cohort.xlsx
"""
import os, glob, argparse, sys, numpy as np
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from plori import core as P
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def safe_sheet(name, prefix, used):
    """Build a unique, Excel-legal (<=31 char) sheet name. openpyxl SILENTLY appends a
    suffix to a duplicate title, so two samples sharing a name (e.g. across categories)
    would yield indistinguishable sheets. Disambiguate deterministically and WARN."""
    s = f"{prefix}{name}"
    for c in '[]:*?/\\':
        s = s.replace(c, "_")
    s = s[:31]
    if s in used:
        i = 1
        while True:
            suf = f"~{i}"
            cand = s[:31 - len(suf)] + suf
            if cand not in used:
                print(f"WARNING: sheet-name collision {s!r} -> {cand!r} "
                      f"(names not unique within 31 chars); rename samples to disambiguate.", flush=True)
                s = cand; break
            i += 1
    used.add(s)
    return s


# (npz key, column title, unit/description)
SUMMARY_COLS = [
    ("name",       "sample",          ""),
    ("cat",        "category",        ""),
    ("mag",        "magnification",   ""),
    ("fps",        "fps",             "frames/s"),
    ("dur",        "duration",        "s"),
    ("nframes",    "frames",          ""),
    ("mpp",        "mpp",             "µm/px"),
    ("area_mm2",   "area",            "mm²"),
    ("opacity",    "opacity",         "%"),
    ("drift_um",   "drift",           "µm"),
    ("bpm_pk",     "BPM",             "beats/min"),
    ("iCV",        "interval CV",     "%"),
    ("cd50_med",   "CD50 (median)",   "ms"),
    ("cd50CV",     "CD50 CV",         "%"),
    ("ct_med",     "CT (median)",     "ms"),
    ("ctCV",       "CT CV",           "%"),
    ("rt_med",     "RT (median)",     "ms"),
    ("rtCV",       "RT CV",           "%"),
    ("amp_med",    "amplitude (median)", "a.u."),
    ("ampCV",      "amplitude CV",    "%"),
    ("max_speed_med",    "max speed (median)",    "µm/s"),   # only present with --with-flow
    ("displacement_med", "displacement (median)", "µm"),     # masked optical flow
    ("strain_med",       "strain (median)",       ""),
    ("k",          "dominant period k", ""),
    ("flags",      "QC flags",        ""),
]

BEAT_COLS = [   # (column title, per-beat array npz key)
    ("interval_ms", "ivl_ms"), ("CD50_ms", "cd50"),
    ("CT_ms", "ct"), ("RT_ms", "rt"), ("amplitude_au", "amp"),
    ("max_speed_umps", "max_speed"), ("displacement_um", "displacement"), ("strain", "strain"),  # --with-flow only
]

README = [
    ("PLoRI results (Excel) — how to read this file", True),
    ("", False),
    ("This file contains beating-analysis results from cardiac organoid videos. Everything is computed from brightness changes alone, with no additional equipment or dyes.", False),
    ("", False),
    ("[Sheet layout]", True),
    ("• Summary : one summary row per sample (all of the metrics below).", False),
    ("• TS_<sample> : the three signal values as a time series, one row per frame (= time). You can plot them directly with Excel charts.", False),
    ("• beats_<sample> : metrics for each individual detected beat.", False),
    ("", False),
    ("[The three signals (TS sheets)]", True),
    ("• PLoRI : a brightness-based signal representing contraction strength. Each peak corresponds to one beat.", False),
    ("• frame_diff : the brightness difference between adjacent frames, |I(t)−I(t−1)|. Reflects the speed of motion.", False),
    ("• masked_flow_umps : masked dense optical-flow speed inside the organoid (µm/s); present only when analysis was run --with-flow.", False),
    ("• dPLoRI_dt : the rate of change of PLoRI. Positive = contracting, negative = relaxing.", False),
    ("", False),
    ("[Summary metrics (Summary sheet)]", True),
    ("• BPM : beats per minute.", False),
    ("• interval CV : irregularity of the beat-to-beat interval (%). Higher means more irregular (arrhythmic tendency).", False),
    ("• CD50 : beat duration at the 50% level (ms). A proxy for calcium-transient duration.", False),
    ("• CT : contraction time (onset → peak, ms). RT : relaxation time (peak → end, ms).", False),
    ("• amplitude : peak height (a.u.). CV : beat-to-beat variability of each metric (%).", False),
    ("• max speed / displacement / strain : optional (only with --with-flow), from masked dense optical flow inside the organoid mask over each beat window. max speed in µm/s; displacement in µm (0.5 x the integral of speed); strain = displacement / organoid size. Blank when --with-flow was not used. Independent of the ContractionWave baseline.", False),
    ("• drift : how far the organoid was displaced during acquisition (µm). Large values warrant QC caution.", False),
    ("• QC flags : automatic quality warnings (e.g. period_lock, amp_var).", False),
    ("", False),
    ("Note: for videos played at a modified speed (e.g. 2x), the time-based values (BPM, CD50, etc.) are off by the same factor.", False),
]


def write_readme(wb):
    ws = wb.active; ws.title = "README"
    ws.column_dimensions["A"].width = 100
    for i, (line, bold) in enumerate(README, 1):
        c = ws.cell(i, 1, line)
        c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--glob", required=True, help="glob of *_ploridata.npz (recursive ** supported)")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    a = ap.parse_args()
    fps_list = sorted(glob.glob(a.glob, recursive=True))
    if not fps_list:
        print(f"no ploridata: {a.glob}"); return

    wb = Workbook()
    write_readme(wb)

    wsS = wb.create_sheet("Summary")
    for j, (_, title, unit) in enumerate(SUMMARY_COLS, 1):
        cell = wsS.cell(1, j, title + (f"\n({unit})" if unit else ""))
        cell.font = Font(bold=True); cell.alignment = Alignment(wrap_text=True, horizontal="center")
        wsS.column_dimensions[get_column_letter(j)].width = 13
    wsS.freeze_panes = "A2"

    used_sheets = set()   # guard against silent openpyxl sheet-name collisions (same name across categories)
    for row, fp in enumerate(fps_list, 2):
        d = np.load(fp, allow_pickle=True)
        name = str(d["name"])
        cat = str(d["cat"]) if "cat" in d.files else ""
        ident = f"{cat}_{name}" if cat else name   # category-qualified so t0h/t24h/... samples get distinct sheets
        # --- Summary row ---
        for j, (key, _, _) in enumerate(SUMMARY_COLS, 1):
            v = d[key] if key in d.files else ""
            if isinstance(v, np.ndarray) and v.ndim == 0:
                v = v.item()
            if isinstance(v, np.floating):
                v = round(float(v), 4)
            elif isinstance(v, np.integer):
                v = int(v)
            elif isinstance(v, np.ndarray):
                v = str(v)
            elif not isinstance(v, (int, float, str)):
                v = str(v)
            wsS.cell(row, j, v)

        # --- Time-series sheet ---
        y = np.asarray(d["plori"], float); T = len(y); fps = float(d["fps"]); dd = np.asarray(d["dd"], float)
        if "pixraw" in d.files and "mask_union" in d.files:
            try:
                fdiff = P.derive(d["pixraw"], d["pix_idx"], d["mask_union"], fps, "ftf")
            except Exception as e:
                fdiff = np.full(T, np.nan); print(f"{name}: frame_diff reconstruction failed {e}")
        else:
            fdiff = np.full(T, np.nan)   # old cache (no per-pixel data): leave column empty
        cols = [("frame", None), ("time_s", None), ("PLoRI", y), ("frame_diff", fdiff)]
        if "flow_speed" in d.files:                                  # masked optical flow speed(t) µm/s (only with --with-flow)
            cols.append(("masked_flow_umps", np.asarray(d["flow_speed"], float)))
        cols.append(("dPLoRI_dt", dd))
        wsT = wb.create_sheet(safe_sheet(ident, "TS_", used_sheets))
        for j, (h, _) in enumerate(cols, 1):
            wsT.cell(1, j, h).font = Font(bold=True)
        wsT.freeze_panes = "A2"
        for i in range(T):
            wsT.cell(i + 2, 1, i)
            wsT.cell(i + 2, 2, round(i / fps, 4))
            for j, (h, arr) in enumerate(cols, 1):
                if arr is None:
                    continue
                v = float(arr[i]) if i < len(arr) else np.nan
                wsT.cell(i + 2, j, round(v, 6) if np.isfinite(v) else "")

        # --- Per-beat sheet ---
        pk = np.asarray(d["pk"]).astype(float)
        ons = np.asarray(d["ons"], float); offs = np.asarray(d["offs"], float)
        wsB = wb.create_sheet(safe_sheet(ident, "beats_", used_sheets))
        bh = ["beat", "peak_time_s"] + [t for t, _ in BEAT_COLS] + ["onset_s", "offset_s"]
        for j, h in enumerate(bh, 1):
            wsB.cell(1, j, h).font = Font(bold=True)
        wsB.freeze_panes = "A2"
        # Per-beat arrays are row-aligned to `pk` ONLY if no beat was skipped. beat_metrics
        # drops beats with A<=0 (so cd50/ct/rt/amp/ons/offs may be shorter) and beat_mechanics
        # drops windows < 2 frames (so max_speed/displacement/strain may be shorter still), each
        # by a DIFFERENT rule. Writing arr[i] against pk-row i would then splice metrics from
        # different beats into one row. So: emit a per-beat column only when its length matches
        # the peak count; otherwise leave it blank and warn (use the Summary medians instead).
        nbeats = len(pk); blanked = []
        beat_arr = {}
        for title, key in BEAT_COLS:
            if key not in d.files:
                beat_arr[key] = None                       # metric absent (e.g. no --with-flow) — blank silently
                continue
            arr = np.asarray(d[key], float)
            if key == "ivl_ms":
                # intervals are inherently n-1 (between consecutive peaks): row i = beat i's forward
                # interval, last row blank. Aligned by construction — accept n or n-1, never warn.
                beat_arr[key] = arr if len(arr) in (nbeats, nbeats - 1) else None
            elif len(arr) == nbeats:
                beat_arr[key] = arr
            else:
                beat_arr[key] = None
            if beat_arr[key] is None and key != "ivl_ms":
                blanked.append(title)
        ons_ok = len(ons) == nbeats; offs_ok = len(offs) == nbeats
        if not ons_ok: blanked.append("onset_s")
        if not offs_ok: blanked.append("offset_s")
        for i in range(nbeats):
            wsB.cell(i + 2, 1, i + 1)
            wsB.cell(i + 2, 2, round(pk[i] / fps, 4))
            for j, (_, key) in enumerate(BEAT_COLS, 3):
                arr = beat_arr[key]
                wsB.cell(i + 2, j, round(float(arr[i]), 4) if (arr is not None and i < len(arr)) else "")
            wsB.cell(i + 2, len(bh) - 1, round(ons[i] / fps, 4) if ons_ok else "")
            wsB.cell(i + 2, len(bh),     round(offs[i] / fps, 4) if offs_ok else "")
        if blanked:
            print(f"WARNING: {name}: {nbeats} peaks but per-beat column(s) {blanked} have a different "
                  f"length (beats skipped in metric/mechanics computation) — left blank to avoid "
                  f"row misalignment; use the Summary medians.", flush=True)
        print(f"[{name}] T={T} beats={len(pk)} sheets: TS_/beats_", flush=True)

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    wb.save(a.out)
    print(f"SAVED {a.out}  ({len(fps_list)} samples, {len(wb.sheetnames)} sheets)", flush=True)


if __name__ == "__main__":
    main()
